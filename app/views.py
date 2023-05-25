import dataclasses
from io import BytesIO
import json
from django.shortcuts import render
from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework.generics import GenericAPIView
from rest_framework.parsers import MultiPartParser
from .serializers import SensorpropertySerializer,SensorExcelSerializer,SensorDataSerializer,SensorpropertySerializers,SensorSerializerrr,uploadExcelserializer,AdvisorySerializer,LayerSerializer,JsonuploadSerializer,SensorlistSerializer,sensorSerializer,SensordataviewSerializer
from rest_framework import status
from rest_framework.response import Response
from .models import Sensor,Sensorproperty
from django.shortcuts import get_object_or_404
from .serializers import SensorSerializerss
from rest_framework import generics
from .serializers import SensorSerializerdd
from rest_framework.exceptions import NotFound
from .models import Sensor,Advisory,Layers
from django.contrib.gis.geos import Point
from django.http import HttpResponse
import pandas as pd
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
# Create your views here.





# sensor post api
class SensorDataCreateded(GenericAPIView):
    serializer_class=SensorDataSerializer
    parser_classes=[MultiPartParser]
    def post(self, request, format=None):
        serializer = SensorDataSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)
 


# sensor property post api
class Sensorpropertyview(GenericAPIView):
    serializer_class=SensorpropertySerializers
    parser_classes=[MultiPartParser]
    def post(self, request,format=None):
     
        serializer = SensorpropertySerializers(data=request.data)
        if serializer.is_valid():
            serializer.save()

            return Response( {'msg':'succfully_Createded'},status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    





# sensor device list
class SensorAPIView(APIView):
    def get(self, request):
             devices = Sensor.objects.all()
             serializer = sensorSerializer(devices, many=True)
             return Response(serializer.data, status=status.HTTP_200_OK)




# device_id wise fetch data
class SensorDataView(APIView):
    def get(self,request,device_id,format=None):
        try:
            sensor_data = Sensor.objects.filter(property__device_id=device_id)
            serializer = SensordataviewSerializer(sensor_data,many=True)
            return Response(serializer.data)
        except Sensor.DoesNotExist:
            return Response({
                'error': 'Device not found'
            }, status=404) 




        
# sensorproperty fectch data_list
class Sensorpropertyviewer(APIView):
    def get(self, request):
             devices = Sensorproperty.objects.all()
             serializer = SensorpropertySerializer(devices, many=True)
             return Response(serializer.data,status=status.HTTP_200_OK)

    


# fectch data sensor with device_id,created_at
class SensorData(APIView):
    serializer_class = SensorSerializerrr

    def get(self, request, device_id, created_at):
        sensor_data = Sensor.objects.filter(property__device_id=device_id,created_at=created_at)
        serializer = self.serializer_class(sensor_data, many=True)
        return Response(serializer.data)
    





# fectch data with sensor_type and created_at and device_id
class SensorListView(generics.ListAPIView):
    serializer_class = SensorSerializerdd
    queryset = Sensor.objects.all()





class SensorList(APIView):
    def get(self, request, sensor_type=None, start_date=None, end_date=None):
        # Retrieve the list of Sensorproperty objects
        sensor_properties = Sensorproperty.objects.all()
        
        # Get a list of unique sensor types
        sensor_types = set([prop.sensor_type for prop in sensor_properties])
      
        
        if not sensor_type:
            # Return the list of sensor types
            return Response(sensor_types)
        
        # Filter the Sensorproperty objects by sensor_type
        sensor_properties = Sensorproperty.objects.filter(sensor_type=sensor_type)
        # print(sensor_properties)
        
        # Filter the Sensor objects by property and created_at between start_date and end_date for each Sensorproperty
        sensors = []
        for sensor_property in sensor_properties:
            sensor_queryset = Sensor.objects.filter(property=sensor_property, 
                                                    created_at__range=(start_date, end_date))
            # print(sensor_queryset)
            sensors.extend(sensor_queryset)
        
        # Check if any Sensor objects were found
        if not sensors:
            return Response({"message": "No Sensor data found for the specified type and date."},
                            status=status.HTTP_404_NOT_FOUND)
        
        # Serialize the data and return a JSON response
        serializer = SensorlistSerializer(sensors, many=True)
        return Response(serializer.data)
    


    



class GeosensorList(APIView):
    def get(self, request, sensor_type=None, start_date=None, end_date=None):
        # Retrieve the list of Sensorproperty objects
        sensor_properties = Sensorproperty.objects.all()
        # Get a list of unique sensor types
        sensor_types = set([prop.sensor_type for prop in sensor_properties])
    
        if not sensor_type:
            # Return the list of sensor types
            return Response(sensor_types)
        
        # Filter the Sensorproperty objects by sensor_type
        sensor_properties = Sensorproperty.objects.filter(sensor_type=sensor_type)
        
        # Filter the Sensor objects by property and created_at between start_date and end_date for each Sensorproperty
        sensors = []
        for sensor_property in sensor_properties:
            sensor_queryset = Sensor.objects.filter(property=sensor_property, 
                                                    created_at__range=(start_date, end_date))
         
        
            sensors.extend(sensor_queryset)
        
        # Check if any Sensor objects were found
        if not sensors:
            return Response({"message": "No Sensor data found for the specified type and date."},
                            status=status.HTTP_404_NOT_FOUND)
        
        # Serialize the data in GeoJSON format
        data = {
            'type': 'FeatureCollection',
            'features': []
        }
        for sensor in sensors:
            point = Point(sensor.property.longitude, sensor.property.latitude)
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': [point.x, point.y]
                },
                'properties': SensorlistSerializer(sensor).data
            }
            data['features'].append(feature)
        
        # Return the GeoJSON response
        return Response(data)
    
 








class GeosensorExcelList(APIView):
    def get(self, request, sensor_type=None, start_date=None, end_date=None):
        # Retrieve the list of Sensorproperty objects
        sensor_properties = Sensorproperty.objects.all()
        
        # Get a list of unique sensor types
        sensor_types = set([prop.sensor_type for prop in sensor_properties])
      
        if not sensor_type:
            # Return the list of sensor types
            return Response(sensor_types)
        
        # Filter the Sensorproperty objects by sensor_type
        sensor_properties = Sensorproperty.objects.filter(sensor_type=sensor_type)
        
        # Filter the Sensor objects by property and created_at between start_date and end_date for each Sensorproperty
        sensors = []
        for sensor_property in sensor_properties:
            sensor_queryset = Sensor.objects.filter(property=sensor_property, created_at__range=(start_date, end_date))
            sensors.extend(sensor_queryset)
        
        # Check if any Sensor objects were found
        if not sensors:
            return Response({"message": "No Sensor data found for the specified type and date."}, status=status.HTTP_404_NOT_FOUND)
        
        # Serialize the data in GeoJSON format
        data = {
            'type': 'FeatureCollection',
            'features': []
        }
        for sensor in sensors:
            point = Point(sensor.property.longitude, sensor.property.latitude)
            feature = {
                'type': 'Feature',
                'geometry': {
                    'type': 'Point',
                    'coordinates': [point.x, point.y]
                },
                'properties': SensorlistSerializer(sensor).data
            }
            data['features'].append(feature)
        
        # Convert GeoJSON data to pandas DataFrame
        df = pd.json_normalize(data['features']).drop(['geometry.type'], axis=1) 
        df['geometry'] = df['geometry.coordinates'].apply(lambda coords: Point(coords[0], coords[1]))
        df.drop(['geometry.coordinates'], axis=1, inplace=True)
      

        # Rename columns
        df.rename(columns={
            "properties.Battery": "Battery",
            "properties.EC_S2": "EC_S2",
            "properties.EC_S1": "EC_S1",
            "properties.Temperature_S1_P": "Temperature_S1_P",
            "properties.Temperature_S2_P": "Temperature_S2_P",
            "properties.Moisture_S1_P":"Moisture_S1_P",
            "properties.Moisture_S2_P":"Moisture_S2_P",
            "properties.device_id":"device_id",
            "properties.created_at": "created_at"
        }, inplace=True)
        
        # Write DataFrame to Excel file
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        # df.drop(['type'], axis=1).to_excel(writer, sheet_name='Sheet1', index=False)
        df.drop(['type'], axis=1).to_excel(writer, sheet_name='Sheet1', index=True)
      
        # df[['Battery', 'EC_S2', 'EC_S1', 'Temperature_S1_P', 'Temperature_S2_P', 'Moisture_S1_P', 'Moisture_S2_P', 'device_id', 'created_at']].to_excel(writer, sheet_name='Sheet1', index=False)

        writer.book.close()  

        # Set the response content type and headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sensor_data.xlsx"'
        
        # Write the Excel file to the response
        output.seek(0)
        response.write(output.getvalue())
        return response







# # # # Advisory

class DumpExcelInsertxlsx(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = uploadExcelserializer
    def post(self, request, format=None):
        # try:
        if 'excel_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."}, status=400)
        excel_file = request.FILES["excel_file"]
        if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:

            return Response({"status": "error","message": "only .xlsx file is supported."},status=400)

        workbook = load_workbook(filename=excel_file)
    
        # print('------220',workbook)
        sheet_name = workbook.sheetnames[0]
        # print(sheet_name)
        worksheet = workbook[sheet_name]
        # print(worksheet)
        data_list = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row): 
                break  
            data = Advisory(crop_name=row[0], Stage=row[1], Agromet_Advisory=row[2])
            data_list.append(data)
        # print(data_list)
        Advisory.objects.bulk_create(data_list)
        return Response({"status": "Success","message": "Successfully Uploaded."})
    








class AdvisoryList(APIView):
    def get(self, request, created_at):
        advisories = Advisory.objects.filter(created_at=created_at)
        if not advisories:
            raise NotFound(detail='No advisories found for the given created_at')
        serializer = AdvisorySerializer(advisories, many=True)
        return Response(serializer.data)


class LayerList(APIView):
    def get(self, request):
        layer_list = Layers.objects.all()
        serializer = LayerSerializer(layer_list, many=True)
        return Response(serializer.data)
    

class LayerPost(GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = LayerSerializer
    def post(self, request):
        serializer = LayerSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)
        




    
# json data upload with device_id 
class SensorUploadView(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = JsonuploadSerializer

    def post(self, request):
        serializer = JsonuploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            # device_id = serializer.validated_data['device_id']

            # Process the uploaded file and save the data to the models
            json_data = file.read().decode('utf-8')

            sensor_data = json.loads(json_data)
            if isinstance(sensor_data, list):
              
                for data in sensor_data:
                    device = data["identifiers"]
                    for id in device:
                        device_id = id["device_ids"]["device_id"]
                        # print(device_id)
                    if "decoded_payload" in data.get("data", {}).get("uplink_message", {}):
                        decoded_payload = data["data"]["uplink_message"]["decoded_payload"]

                        Battery = decoded_payload.get('Battery')
                        Temperature_S1_P = decoded_payload.get('Temperature_S1_P')
                        Temperature_S2_P = decoded_payload.get('Temperature_S2_P')
                        EC_S1 = decoded_payload.get('EC_S1')
                        EC_S2 = decoded_payload.get('EC_S2')
                        Moisture_S1_P = decoded_payload.get('Moisture_S1_P')
                        Moisture_S2_P = decoded_payload.get('Moisture_S2_P')

                        if all(value is not None for value in [Battery, EC_S1, EC_S2, Moisture_S1_P, Moisture_S2_P, Temperature_S1_P, Temperature_S2_P]):
                            try:

                                sensor_property = Sensorproperty.objects.get(device_id=device_id).pk
                                
                                # print(sensor_property.device_id)
                            except:
                                return Response ({"message" : "Devide ID not Found"})
                            create_data = Sensor.objects.create(
                            Battery=float(Battery),
                            EC_S1=float(EC_S1),
                            EC_S2=float(EC_S2),
                            Moisture_S1_P=float(Moisture_S1_P),
                            Moisture_S2_P=float(Moisture_S2_P),
                            Temperature_S1_P=float(Temperature_S1_P),
                            Temperature_S2_P=float(Temperature_S2_P),
                            property_id = sensor_property )

                    else:
                        continue

            return Response({'message': 'Data saved successfully.'}, status=status.HTTP_201_CREATED)                      
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        










